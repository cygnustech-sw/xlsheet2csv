"""Core workbook discovery, validation, conversion and manifest logic."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import struct
import unicodedata
import uuid
import zipfile
from collections.abc import Iterable
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import load_workbook

from . import __version__

TOOL_NAME = "xlsheet2csv"
SUPPORTED_SUFFIXES = {".xlsx"}
WINDOWS_RESERVED_NAMES = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    *(f"COM{number}" for number in range(1, 10)),
    *(f"LPT{number}" for number in range(1, 10)),
}


class ConversionError(Exception):
    """Expected conversion failure suitable for a per-workbook result."""


class InputLimitError(ConversionError):
    """Workbook exceeded a configured resource or archive limit."""


class ExistingOutputError(ConversionError):
    """A deterministic output directory already exists."""


@dataclass(frozen=True)
class ConversionPolicy:
    encoding: str = "utf-8"
    delimiter: str = ","
    formulas: str = "values"
    formula_safety: str = "escape"
    hidden_sheets: str = "include"
    trailing_empty_cells: str = "omit"
    line_ending: str = "LF"

    def validate(self) -> None:
        if self.encoding not in {"utf-8", "utf-8-sig"}:
            raise ValueError(f"Unsupported encoding: {self.encoding}")
        if len(self.delimiter) != 1:
            raise ValueError("Delimiter must be exactly one character.")
        if self.formulas not in {"values", "formulas"}:
            raise ValueError(f"Unsupported formula policy: {self.formulas}")
        if self.formula_safety not in {"escape", "preserve"}:
            raise ValueError(f"Unsupported formula safety policy: {self.formula_safety}")
        if self.hidden_sheets not in {"include", "exclude"}:
            raise ValueError(f"Unsupported hidden-sheet policy: {self.hidden_sheets}")


@dataclass(frozen=True)
class InputLimits:
    max_file_bytes: int = 512 * 1024 * 1024
    max_expanded_bytes: int = 2 * 1024 * 1024 * 1024
    max_compression_ratio: float = 250.0
    max_archive_members: int = 20_000
    max_central_directory_bytes: int = 64 * 1024 * 1024
    max_sheets: int = 1024
    max_rows_per_sheet: int = 1_048_576
    max_columns_per_sheet: int = 16_384

    def validate(self) -> None:
        for name, value in asdict(self).items():
            if value <= 0:
                raise ValueError(f"{name} must be greater than zero.")


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_stream(handle: BinaryIO, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    original_position = handle.tell()
    try:
        handle.seek(0)
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    finally:
        handle.seek(original_position)
    return digest.hexdigest()


def safe_component(value: str, fallback: str = "unnamed", max_length: int = 80) -> str:
    normalised = unicodedata.normalize("NFKC", value)
    safe = "".join("_" if character in '\\/:*?"<>|' or ord(character) < 32 else character for character in normalised)
    safe = safe.strip().rstrip(".")
    if not safe:
        safe = fallback
    if safe.split(".", 1)[0].upper() in WINDOWS_RESERVED_NAMES:
        safe = f"_{safe}"
    return safe[:max_length].rstrip(" .") or fallback


def _is_within(path: Path, parent: Path) -> bool:
    try:
        path.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def discover_workbooks(source: Path, recurse: bool, output_root: Path) -> tuple[list[Path], Path]:
    source = source.expanduser().resolve()
    if not source.exists():
        raise ConversionError(f"Source path not found: {source}")

    if source.is_file():
        if source.suffix.casefold() not in SUPPORTED_SUFFIXES:
            raise ConversionError("Source file must have an .xlsx extension.")
        return [source], source.parent

    iterator: Iterable[Path] = source.rglob("*") if recurse else source.iterdir()
    discovered: set[Path] = set()
    for candidate in iterator:
        if (
            not candidate.is_file()
            or candidate.suffix.casefold() not in SUPPORTED_SUFFIXES
            or candidate.name.startswith("~$")
        ):
            continue
        resolved = candidate.resolve()
        if not _is_within(resolved, source) or _is_within(resolved, output_root):
            continue
        discovered.add(resolved)
    workbooks = sorted(discovered, key=lambda candidate: (str(candidate).casefold(), str(candidate)))
    if not workbooks:
        raise ConversionError("No .xlsx files were found.")
    return workbooks, source


def _preflight_central_directory(handle: BinaryIO, file_size: int, limits: InputLimits) -> int:
    """Count central-directory entries without materialising ZipInfo objects."""
    end_record = struct.Struct("<4s4H2LH")
    tail_length = min(file_size, 22 + 65_535)
    handle.seek(-tail_length, os.SEEK_END)
    tail = handle.read(tail_length)
    end_index = tail.rfind(b"PK\x05\x06")
    if end_index < 0 or len(tail) - end_index < end_record.size:
        raise ConversionError("File does not contain a valid ZIP end record.")

    (
        _,
        disk_number,
        directory_disk,
        entries_on_disk,
        declared_entries,
        directory_size,
        directory_offset,
        comment_length,
    ) = end_record.unpack_from(tail, end_index)
    end_offset = file_size - tail_length + end_index
    if end_offset + end_record.size + comment_length != file_size:
        raise ConversionError("ZIP end record or comment length is inconsistent.")
    if disk_number != 0 or directory_disk != 0 or entries_on_disk != declared_entries:
        raise ConversionError("Multi-disk ZIP archives are not supported.")
    if declared_entries == 0xFFFF or directory_size == 0xFFFFFFFF or directory_offset == 0xFFFFFFFF:
        raise ConversionError("ZIP64 central directories are not supported within the XLSX input limits.")
    if declared_entries > limits.max_archive_members:
        raise InputLimitError(
            f"Workbook archive has {declared_entries} members; limit is {limits.max_archive_members}."
        )
    if directory_size > limits.max_central_directory_bytes:
        raise InputLimitError(
            "Workbook central directory is "
            f"{directory_size} bytes; limit is {limits.max_central_directory_bytes} bytes."
        )
    if directory_offset + directory_size > end_offset:
        raise ConversionError("ZIP central-directory bounds are inconsistent.")

    handle.seek(directory_offset)
    consumed = 0
    counted_entries = 0
    while consumed < directory_size:
        header = handle.read(46)
        if len(header) != 46 or header[:4] != b"PK\x01\x02":
            raise ConversionError("ZIP central directory contains an invalid entry header.")
        variable_length = sum(struct.unpack_from("<3H", header, 28))
        entry_length = 46 + variable_length
        if consumed + entry_length > directory_size:
            raise ConversionError("ZIP central-directory entry exceeds its declared bounds.")
        handle.seek(variable_length, os.SEEK_CUR)
        consumed += entry_length
        counted_entries += 1
        if counted_entries > limits.max_archive_members:
            raise InputLimitError(f"Workbook archive has more than {limits.max_archive_members} members.")

    if counted_entries != declared_entries:
        raise ConversionError("ZIP central-directory member count is inconsistent.")
    return counted_entries


def validate_workbook_archive(
    source: Path | BinaryIO,
    limits: InputLimits,
) -> dict[str, int | float]:
    limits.validate()
    if isinstance(source, Path):
        with source.open("rb") as handle:
            return validate_workbook_archive(handle, limits)

    handle = source
    file_size = os.fstat(handle.fileno()).st_size
    if file_size > limits.max_file_bytes:
        raise InputLimitError(f"Workbook is {file_size} bytes; limit is {limits.max_file_bytes} bytes.")
    member_count = _preflight_central_directory(handle, file_size, limits)

    handle.seek(0)
    with zipfile.ZipFile(handle) as archive:
        members = archive.infolist()
        if len(members) != member_count:
            raise ConversionError("ZIP member count changed during validation.")
        if any(member.flag_bits & 0x1 for member in members):
            raise ConversionError("Encrypted XLSX archives are not supported.")
        expanded_size = sum(member.file_size for member in members)
        compressed_size = sum(member.compress_size for member in members)
    handle.seek(0)

    if expanded_size > limits.max_expanded_bytes:
        raise InputLimitError(
            f"Expanded workbook is {expanded_size} bytes; limit is {limits.max_expanded_bytes} bytes."
        )
    compression_ratio = expanded_size / max(compressed_size, 1)
    if compression_ratio > limits.max_compression_ratio:
        raise InputLimitError(
            f"Workbook compression ratio is {compression_ratio:.1f}; limit is {limits.max_compression_ratio:.1f}."
        )
    return {
        "file_bytes": file_size,
        "expanded_bytes": expanded_size,
        "compression_ratio": round(compression_ratio, 2),
        "archive_members": member_count,
    }


def _formula_safe_text(value: str, policy: ConversionPolicy) -> str:
    if policy.formula_safety == "escape" and value.lstrip().startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def serialise_cell(value: Any, policy: ConversionPolicy) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return _formula_safe_text(value, policy)
    if isinstance(value, datetime):
        return value.isoformat(sep="T")
    if isinstance(value, (date, time)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


def _selected_sheet(name: str, state: str, include: set[str], exclude: set[str], policy: ConversionPolicy) -> bool:
    folded = name.casefold()
    if include and folded not in include:
        return False
    if folded in exclude:
        return False
    return not (policy.hidden_sheets == "exclude" and state != "visible")


def _workbook_output_name(path: Path, source_root: Path, source_hash: str) -> str:
    relative = path.relative_to(source_root).as_posix()
    path_hash = hashlib.sha256(relative.encode("utf-8")).hexdigest()[:8]
    return f"{safe_component(path.stem)}--{source_hash[:12]}-{path_hash}"


def convert_workbook(
    workbook_path: Path,
    source_root: Path,
    output_root: Path,
    policy: ConversionPolicy,
    limits: InputLimits,
    include_sheets: Iterable[str] = (),
    exclude_sheets: Iterable[str] = (),
) -> dict[str, Any]:
    policy.validate()
    relative_source = workbook_path.relative_to(source_root).as_posix()
    include = {name.casefold() for name in include_sheets}
    exclude = {name.casefold() for name in exclude_sheets}
    source_handle = workbook_path.open("rb")
    workbook = None
    stage_path = None

    try:
        archive = validate_workbook_archive(source_handle, limits)
        source_hash = sha256_stream(source_handle)
        output_name = _workbook_output_name(workbook_path, source_root, source_hash)
        final_path = output_root / output_name
        if final_path.exists():
            raise ExistingOutputError(f"Output already exists and was not changed: {final_path.name}")

        stage_path = output_root / f".xlsheet2csv-stage-{uuid.uuid4().hex}"
        stage_path.mkdir(parents=False, exist_ok=False)
        source_handle.seek(0)
        workbook = load_workbook(
            source_handle,
            read_only=True,
            data_only=policy.formulas == "values",
            keep_links=False,
        )
        if len(workbook.sheetnames) > limits.max_sheets:
            raise InputLimitError(f"Workbook has {len(workbook.sheetnames)} sheets; limit is {limits.max_sheets}.")

        outputs: list[dict[str, Any]] = []
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            if not _selected_sheet(worksheet.title, worksheet.sheet_state, include, exclude, policy):
                continue
            dimension = worksheet.calculate_dimension(force=True)
            maximum_column = worksheet.max_column or 0
            if maximum_column > limits.max_columns_per_sheet:
                raise InputLimitError(
                    f"Sheet {worksheet.title!r} has {maximum_column} columns; "
                    f"limit is {limits.max_columns_per_sheet}."
                )

            csv_name = f"{sheet_index:03d}--{safe_component(worksheet.title, fallback='sheet')}.csv"
            csv_path = stage_path / csv_name
            row_count = 0
            max_columns_written = 0
            nominal_empty = dimension == "A1:A1"

            with csv_path.open("w", encoding=policy.encoding, newline="") as csv_handle:
                writer = csv.writer(csv_handle, delimiter=policy.delimiter, lineterminator="\n")
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                    if row_index > limits.max_rows_per_sheet:
                        raise InputLimitError(
                            f"Sheet {worksheet.title!r} exceeds the {limits.max_rows_per_sheet}-row limit."
                        )
                    values = [serialise_cell(value, policy) for value in row]
                    while values and values[-1] == "":
                        values.pop()
                    if nominal_empty and row_index == 1 and not values:
                        continue
                    writer.writerow(values)
                    row_count += 1
                    max_columns_written = max(max_columns_written, len(values))

            outputs.append(
                {
                    "sheet_index": sheet_index,
                    "sheet_name": worksheet.title,
                    "sheet_state": worksheet.sheet_state,
                    "csv_file": csv_name,
                    "rows": row_count,
                    "columns": max_columns_written,
                    "sha256": sha256_file(csv_path),
                    "bytes": csv_path.stat().st_size,
                }
            )

        if not outputs:
            raise ConversionError("No worksheets matched the selected filters and hidden-sheet policy.")

        manifest = {
            "schema_version": "1.0",
            "tool": {"name": TOOL_NAME, "version": __version__},
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "source": {
                "relative_path": relative_source,
                "sha256": source_hash,
                **archive,
            },
            "conversion_policy": asdict(policy),
            "limits": asdict(limits),
            "worksheets": outputs,
            "notes": [
                "Cell display formatting is not applied; underlying workbook values are serialised.",
                "Formula value mode uses cached workbook values and does not calculate formulas.",
                "A leading apostrophe is added to formula-like text when formula_safety is escape.",
            ],
        }
        manifest_path = stage_path / "manifest.json"
        manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
        os.replace(stage_path, final_path)
        return {
            "source_relative_path": relative_source,
            "source_sha256": source_hash,
            "output_directory": final_path.name,
            "manifest": f"{final_path.name}/manifest.json",
            "worksheet_count": len(outputs),
        }
    except Exception:
        if stage_path is not None and stage_path.exists():
            shutil.rmtree(stage_path, ignore_errors=True)
        raise
    finally:
        if workbook is not None:
            workbook.close()
        source_handle.close()


def _write_run_manifest(output_root: Path, manifest: dict[str, Any]) -> Path:
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S%fZ")
    name = f"run--{stamp}-{uuid.uuid4().hex[:8]}.json"
    final_path = output_root / name
    temporary_path = output_root / f".{name}.tmp"
    manifest["run_manifest"] = name
    temporary_path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    os.replace(temporary_path, final_path)
    return final_path


def convert_path(
    source: Path,
    output_root: Path,
    recurse: bool = False,
    policy: ConversionPolicy | None = None,
    limits: InputLimits | None = None,
    include_sheets: Iterable[str] = (),
    exclude_sheets: Iterable[str] = (),
) -> tuple[dict[str, Any], int, Path]:
    policy = policy or ConversionPolicy()
    limits = limits or InputLimits()
    output_root = output_root.expanduser().resolve()
    workbooks, source_root = discover_workbooks(source, recurse, output_root)
    output_root.mkdir(parents=True, exist_ok=True)
    started_at = datetime.now(timezone.utc)
    successes: list[dict[str, Any]] = []
    failures: list[dict[str, str]] = []

    for workbook in workbooks:
        relative = workbook.relative_to(source_root).as_posix()
        try:
            successes.append(
                convert_workbook(
                    workbook,
                    source_root,
                    output_root,
                    policy,
                    limits,
                    include_sheets,
                    exclude_sheets,
                )
            )
        except Exception as error:
            failures.append(
                {
                    "source_relative_path": relative,
                    "error_type": type(error).__name__,
                    "message": str(error),
                }
            )

    exit_code = 0 if not failures else (2 if successes else 1)
    run_manifest: dict[str, Any] = {
        "schema_version": "1.0",
        "run_id": uuid.uuid4().hex,
        "tool": {"name": TOOL_NAME, "version": __version__},
        "started_at_utc": started_at.isoformat(),
        "finished_at_utc": datetime.now(timezone.utc).isoformat(),
        "status": "succeeded" if exit_code == 0 else ("partial" if exit_code == 2 else "failed"),
        "input_count": len(workbooks),
        "success_count": len(successes),
        "failure_count": len(failures),
        "conversion_policy": asdict(policy),
        "successes": successes,
        "failures": failures,
    }
    manifest_path = _write_run_manifest(output_root, run_manifest)
    return run_manifest, exit_code, manifest_path
