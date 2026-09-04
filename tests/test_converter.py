from __future__ import annotations

import csv
import json
import os
import tempfile
import unittest
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook

from xlsheet2csv.converter import (
    ConversionError,
    ConversionPolicy,
    InputLimitError,
    InputLimits,
    convert_path,
    serialise_cell,
    sha256_file,
    sha256_stream,
    validate_workbook_archive,
)


def write_workbook(path: Path, marker: str = "value", leading_zero: str = "00123") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["identifier", "marker"])
    worksheet.append([leading_zero, marker])
    workbook.save(path)
    workbook.close()


class ConverterTests(unittest.TestCase):
    def test_recursive_same_basename_does_not_collide_and_preserves_text(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            (root / "input" / "a").mkdir(parents=True)
            (root / "input" / "b").mkdir(parents=True)
            write_workbook(root / "input" / "a" / "report.xlsx", "SAME")
            write_workbook(root / "input" / "b" / "report.xlsx", "SAME")

            manifest, exit_code, _ = convert_path(root / "input", root / "output", recurse=True)

            self.assertEqual(exit_code, 0)
            self.assertEqual(manifest["success_count"], 2)
            directories = {result["output_directory"] for result in manifest["successes"]}
            self.assertEqual(len(directories), 2)
            rows = []
            for directory in directories:
                with next((root / "output" / directory).glob("*.csv")).open(newline="", encoding="utf-8") as handle:
                    rows.extend(csv.reader(handle))
            self.assertEqual(rows.count(["00123", "SAME"]), 2)

    def test_sanitised_sheet_names_are_distinguished_by_index(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook = Workbook()
            workbook.active.title = "A<B"
            workbook.active.append(["left"])
            workbook.create_sheet("A>B").append(["right"])
            workbook.save(root / "collision.xlsx")
            workbook.close()

            manifest, exit_code, _ = convert_path(root / "collision.xlsx", root / "output")

            self.assertEqual(exit_code, 0)
            output = root / "output" / manifest["successes"][0]["output_directory"]
            names = sorted(path.name for path in output.glob("*.csv"))
            self.assertEqual(names, ["001--A_B.csv", "002--A_B.csv"])

    def test_existing_output_is_not_overwritten(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            write_workbook(root / "report.xlsx")
            first, first_exit, _ = convert_path(root / "report.xlsx", root / "output")
            output = root / "output" / first["successes"][0]["output_directory"]
            csv_path = next(output.glob("*.csv"))
            before = csv_path.read_bytes()

            second, second_exit, _ = convert_path(root / "report.xlsx", root / "output")

            self.assertEqual(first_exit, 0)
            self.assertEqual(second_exit, 1)
            self.assertEqual(second["failures"][0]["error_type"], "ExistingOutputError")
            self.assertEqual(csv_path.read_bytes(), before)

    def test_bad_workbook_produces_partial_batch_result(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            (root / "input").mkdir()
            (root / "input" / "broken.xlsx").write_text("not a workbook", encoding="utf-8")
            write_workbook(root / "input" / "valid.xlsx")

            manifest, exit_code, run_path = convert_path(root / "input", root / "output")

            self.assertEqual(exit_code, 2)
            self.assertEqual(manifest["success_count"], 1)
            self.assertEqual(manifest["failure_count"], 1)
            saved_run = json.loads(run_path.read_text(encoding="utf-8"))
            self.assertEqual(saved_run["status"], "partial")
            self.assertEqual(saved_run["run_manifest"], run_path.name)

    def test_sheet_filters_are_case_insensitive(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            write_workbook(root / "report.xlsx")

            manifest, exit_code, _ = convert_path(
                root / "report.xlsx",
                root / "output",
                include_sheets=["data"],
            )

            self.assertEqual(exit_code, 0)
            self.assertEqual(manifest["successes"][0]["worksheet_count"], 1)

    def test_hidden_sheet_policy_is_explicit(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook = Workbook()
            workbook.active.title = "Visible"
            hidden = workbook.create_sheet("Hidden")
            hidden.sheet_state = "hidden"
            workbook.save(root / "report.xlsx")
            workbook.close()

            manifest, exit_code, _ = convert_path(
                root / "report.xlsx",
                root / "output",
                policy=ConversionPolicy(hidden_sheets="exclude"),
            )

            self.assertEqual(exit_code, 0)
            self.assertEqual(manifest["successes"][0]["worksheet_count"], 1)

    def test_formula_like_text_is_escaped_by_default(self) -> None:
        self.assertEqual(serialise_cell("+441234567", ConversionPolicy()), "'+441234567")
        self.assertEqual(
            serialise_cell("+441234567", ConversionPolicy(formula_safety="preserve")),
            "+441234567",
        )

    def test_dates_and_booleans_use_stable_text(self) -> None:
        policy = ConversionPolicy()
        self.assertEqual(serialise_cell(date(2026, 9, 4), policy), "2026-09-04")
        self.assertEqual(serialise_cell(datetime(2026, 9, 4, 12, 30, 5), policy), "2026-09-04T12:30:05")
        self.assertEqual(serialise_cell(True, policy), "true")

    def test_small_input_limit_fails_without_output_directory(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            write_workbook(root / "report.xlsx")
            limits = InputLimits(max_file_bytes=1)

            manifest, exit_code, _ = convert_path(root / "report.xlsx", root / "output", limits=limits)

            self.assertEqual(exit_code, 1)
            self.assertEqual(manifest["failures"][0]["error_type"], "InputLimitError")
            self.assertFalse(any(path.is_dir() for path in (root / "output").iterdir()))

    def test_directory_discovery_does_not_follow_external_symlink(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            (root / "input").mkdir()
            write_workbook(root / "outside.xlsx")
            try:
                (root / "input" / "linked.xlsx").symlink_to(root / "outside.xlsx")
            except (NotImplementedError, OSError):
                self.skipTest("File symlinks are unavailable on this platform.")

            with self.assertRaisesRegex(ConversionError, "No .xlsx files were found"):
                convert_path(root / "input", root / "output", recurse=True)

    def test_archive_resource_limits_are_enforced(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            workbook_path = Path(temporary) / "report.xlsx"
            write_workbook(workbook_path)
            limits_and_messages = (
                (InputLimits(max_expanded_bytes=1), "Expanded workbook"),
                (InputLimits(max_compression_ratio=0.1), "compression ratio"),
                (InputLimits(max_archive_members=1), "archive has"),
                (InputLimits(max_central_directory_bytes=1), "central directory"),
            )

            for limits, message in limits_and_messages:
                with self.subTest(message=message):
                    with self.assertRaisesRegex(InputLimitError, message):
                        validate_workbook_archive(workbook_path, limits)

    def test_invalid_source_does_not_create_output_root(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            output = root / "output"

            with self.assertRaisesRegex(ConversionError, "Source path not found"):
                convert_path(root / "missing", output)

            self.assertFalse(output.exists())

    def test_atomic_source_replacement_cannot_change_parsed_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "report.xlsx"
            replacement = root / "replacement.xlsx"
            write_workbook(source, marker="ORIGINAL")
            write_workbook(replacement, marker="REPLACEMENT")
            original_hash = sha256_file(source)

            def replace_after_hash(handle):
                result = sha256_stream(handle)
                try:
                    os.replace(replacement, source)
                except PermissionError:
                    self.skipTest("The platform prevents replacement of an open input file.")
                return result

            with patch("xlsheet2csv.converter.sha256_stream", side_effect=replace_after_hash):
                manifest, exit_code, _ = convert_path(source, root / "output")

            self.assertEqual(exit_code, 0)
            output = root / "output" / manifest["successes"][0]["output_directory"]
            workbook_manifest = json.loads((output / "manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(workbook_manifest["source"]["sha256"], original_hash)
            with next(output.glob("*.csv")).open(newline="", encoding="utf-8") as handle:
                rows = list(csv.reader(handle))
            self.assertIn(["00123", "ORIGINAL"], rows)
            self.assertNotIn(["00123", "REPLACEMENT"], rows)


if __name__ == "__main__":
    unittest.main()
